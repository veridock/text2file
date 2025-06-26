"""Excel file generator module."""

import io
import logging
from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseGenerator

logger = logging.getLogger(__name__)


class ExcelGenerator(BaseGenerator):
    """Generator for Excel files (.xlsx, .xls)."""

    @classmethod
    def generate(
        cls,
        content: str,
        output_path: Union[str, Path],
        **kwargs: Any,
    ) -> Path:
        """Generate an Excel file from the given content.
        
        Args:
            content: The content to convert to Excel. Should be in CSV format.
            output_path: Path where the Excel file will be saved.
            **kwargs: Additional keyword arguments:
                - sheet_name: Name of the worksheet (default: "Sheet1")
                - auto_adjust: Whether to auto-adjust column widths (default: True)
                
        Returns:
            Path to the generated Excel file.
            
        Raises:
            ValueError: If the content cannot be parsed as CSV
            IOError: If there's an error writing the Excel file
        """
        sheet_name = kwargs.get('sheet_name', 'Sheet1')
        auto_adjust = kwargs.get('auto_adjust', True)
        
        output_path = Path(output_path)
        
        try:
            # Parse the CSV content
            rows = _parse_csv_content(content)
            
            # Create a new workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Add data to worksheet
            for row_idx, row in enumerate(rows, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value.strip())
            
            # Auto-adjust column widths if requested
            if auto_adjust and rows:
                _auto_adjust_columns(ws)
            
            # Save the workbook
            wb.save(str(output_path))
            logger.debug(f"Generated Excel file: {output_path}")
            
            return output_path
            
        except Exception as e:
            # Clean up partially created file
            if output_path.exists():
                output_path.unlink()
            raise IOError(f"Error generating Excel file: {str(e)}") from e

def _parse_csv_content(content: str) -> list[list[str]]:
    """Parse CSV-like content into a 2D list of strings.
    
    Args:
        content: CSV content as a string
        
    Returns:
        List of rows, where each row is a list of cell values
        
    Raises:
        ValueError: If the content cannot be parsed as CSV
    """
    try:
        return [line.split(',') for line in content.strip().split('\n') if line.strip()]
    except Exception as e:
        raise ValueError(f"Failed to parse CSV content: {str(e)}") from e

def _auto_adjust_columns(worksheet: Worksheet) -> None:
    """Auto-adjust column widths based on content.
    
    Args:
        worksheet: The worksheet to adjust
    """
    for column_cells in worksheet.columns:
        if not column_cells:
            continue
            
        max_length = 0
        try:
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    continue
            
            if max_length > 0:
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = min(adjusted_width, 50)
        except Exception:
            continue

# Create an instance of the Excel generator
excel_generator = ExcelGenerator()

# Register this generator
def register() -> None:
    """Register the Excel file generator."""
    from .registration import register_generator
    
    @register_generator(["xlsx", "xls"])
    def excel_generator_wrapper(
        content: str,
        output_path: Union[str, Path],
        **kwargs: Any,
    ) -> Path:
        """Generate an Excel file from the given content.
        
        Args:
            content: The content to convert to Excel.
            output_path: Path where the Excel file will be saved.
            **kwargs: Additional keyword arguments for ExcelGenerator.generate().
            
        Returns:
            Path to the generated Excel file.
        """
        return ExcelGenerator.generate(content, output_path, **kwargs)

#!/usr/bin/env python3
from openpyxl import load_workbook

def check_excel_file(filepath):
    try:
        print(f"Checking Excel file: {filepath}")
        wb = load_workbook(filename=filepath)
        print(f"Sheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\nSheet: {sheet_name}")
            ws = wb[sheet_name]
            print("Data:")
            for row in ws.iter_rows(values_only=True):
                print(row)
                
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        check_excel_file(sys.argv[1])
    else:
        print("Please provide the path to an Excel file")
